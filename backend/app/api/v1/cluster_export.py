"""
Cluster data export to Excel.
"""
import json
import logging
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.database import get_db
from app.models.cluster import (
    Cluster, Node, Upstream, UpstreamTarget, Route, RoutePlugin,
    PluginConfig, GlobalRule, PluginMetadata, StreamProxy,
)
from app.models.static_resource import StaticResource
from app.models.ssl import SslCertificate
from app.api.v1.clusters import get_current_user
from app.services.edge_sync import get_or_404

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/clusters", tags=["clusters"])


def _sanitize_filename(name):
    import re
    return re.sub(r'[\\/:*?"<>|#]', '_', name).strip()


def _fmt_json(val):
    if not val:
        return ""
    try:
        parsed = json.loads(val)
        return json.dumps(parsed, ensure_ascii=False, indent=2)
    except (json.JSONDecodeError, TypeError):
        return val or ""


def _fmt_dt(dt):
    if dt is None:
        return ""
    return dt.isoformat()


async def _query_all_data(cluster_id, db):
    import asyncio

    async def _fetch_all(model, **filters):
        q = select(model)
        for k, v in filters.items():
            q = q.where(getattr(model, k) == v)
        r = await db.execute(q)
        return list(r.scalars().all())

    async def _fetch_targets(upstream_ids):
        if not upstream_ids:
            return {}
        q = select(UpstreamTarget).where(
            UpstreamTarget.upstream_id.in_(upstream_ids))
        r = await db.execute(q)
        m = {}
        for t in r.scalars().all():
            m.setdefault(t.upstream_id, []).append(t)
        return m

    async def _fetch_plugins(route_ids):
        if not route_ids:
            return {}
        q = select(RoutePlugin).where(
            RoutePlugin.route_id.in_(route_ids))
        r = await db.execute(q)
        m = {}
        for p in r.scalars().all():
            m.setdefault(p.route_id, []).append(p)
        return m

    cluster = await get_or_404(db, Cluster, id=cluster_id,
                                detail="集群不存在")

    nodes, upstreams, routes, pcs, grs, pms, sps, srs, ssls = \
        await asyncio.gather(
            _fetch_all(Node, cluster_id=cluster_id),
            _fetch_all(Upstream, cluster_id=cluster_id),
            _fetch_all(Route, cluster_id=cluster_id),
            _fetch_all(PluginConfig, cluster_id=cluster_id),
            _fetch_all(GlobalRule, cluster_id=cluster_id),
            _fetch_all(PluginMetadata, cluster_id=cluster_id),
            _fetch_all(StreamProxy, cluster_id=cluster_id),
            _fetch_all(StaticResource, cluster_id=cluster_id),
            _fetch_all(SslCertificate, cluster_id=cluster_id),
        )

    up_ids = [u.id for u in upstreams]
    rt_ids = [r.id for r in routes]
    targets_map, plugins_map = await asyncio.gather(
        _fetch_targets(up_ids), _fetch_plugins(rt_ids),
    )

    up_name_map = {u.id: u.name for u in upstreams}
    rt_name_map = {r.id: r.name for r in routes}
    pc_uuid_map = {p.edge_uuid: (p.id, p.name) for p in pcs if p.edge_uuid}

    return {
        "cluster": cluster,
        "nodes": nodes,
        "upstreams": upstreams,
        "targets_map": targets_map,
        "routes": routes,
        "plugins_map": plugins_map,
        "up_name_map": up_name_map,
        "up_id_map": {u.id: u.id for u in upstreams},
        "rt_name_map": rt_name_map,
        "rt_id_map": {r.id: r.id for r in routes},
        "pc_uuid_map": pc_uuid_map,
        "plugin_configs": pcs,
        "global_rules": grs,
        "plugin_metadatas": pms,
        "stream_proxies": sps,
        "static_resources": srs,
        "ssl_certificates": ssls,
    }


def _build_workbook(data):
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        wb.remove(ws)

    LINK_FONT = Font(color="0563C1", underline="single")

    def write_sheet(sheet_name, headers, rows, link_cols=None):
        """Write a sheet. link_cols is a set of 0-based column indices
        that should be rendered as internal hyperlinks.
        rows[i][col] is expected to be (display_value, target_sheet, target_cell).
        """
        ws = wb.create_sheet(title=sheet_name)
        bold = Font(bold=True)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = bold
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                if link_cols and (ci - 1) in link_cols and isinstance(val, tuple):
                    display, target_sheet, target_cell = val
                    cell = ws.cell(row=ri, column=ci, value=display)
                    cell.hyperlink = f"#{target_sheet}!{target_cell}"
                    cell.font = LINK_FONT
                else:
                    cell = ws.cell(row=ri, column=ci, value=val)
        for ci in range(1, len(headers) + 1):
            ml = len(str(headers[ci - 1] or ""))
            for ri in range(2, min(len(rows) + 2, 50)):
                cv = ws.cell(row=ri, column=ci).value
                if cv:
                    ml = max(ml, min(len(str(cv)), 60))
            ws.column_dimensions[get_column_letter(ci)].width = ml + 3
        return ws

    # ── Pre-compute row maps for FK targets ──────────────────────────
    # Data rows start at row 2 (row 1 = header). Index in the list + 2 = Excel row.
    up_row = {data["upstreams"][i].id: i + 2 for i in range(len(data["upstreams"]))}
    rt_row = {data["routes"][i].id: i + 2 for i in range(len(data["routes"]))}
    pc_row = {}
    for i, p in enumerate(data["plugin_configs"]):
        if p.edge_uuid:
            pc_row[p.edge_uuid] = i + 2

    c = data["cluster"]

    # 1. 集群信息 (no admin_key)
    write_sheet("集群信息", [
        "集群标识", "显示名称", "管理地址", "描述", "分组", "状态",
        "创建时间", "更新时间",
    ], [[
        c.name, c.display_name or "", c.admin_url or "", c.description or "",
        c.group_name or "", "启用" if c.status == 1 else "禁用",
        _fmt_dt(c.created_at), _fmt_dt(c.updated_at),
    ]])

    # 2. 集群节点
    nr = []
    for n in data["nodes"]:
        nr.append([
            n.id, n.ip, n.service_port, n.management_port, n.edge_path,
            n.edge_install_path or "",
            "在线" if n.status == 1 else "离线",
            _fmt_dt(n.created_at),
        ])
    write_sheet("集群节点", [
        "ID", "IP", "服务端口", "管理端口", "Edge 路径", "安装路径", "状态", "创建时间",
    ], nr)

    # 3. 上游服务
    tm = data["targets_map"]
    ur = []
    for u in data["upstreams"]:
        targets = tm.get(u.id, [])
        ts = "; ".join(
            f"{t.target}(权重{t.weight})" for t in targets) if targets else "（无）"
        ur.append([
            u.id, u.name, u.load_balance or "", u.scheme or "",
            u.pass_host or "", u.upstream_host or "",
            _fmt_json(u.timeout), u.retries or "", u.retry_timeout or "",
            _fmt_json(u.checks), _fmt_json(u.keepalive_pool),
            ts, u.description or "", _fmt_dt(u.created_at),
        ])
    write_sheet("上游服务", [
        "ID", "名称", "负载均衡", "协议", "Pass Host", "上游 Host",
        "超时", "重试次数", "重试超时", "健康检查", "连接池",
        "目标节点", "描述", "创建时间",
    ], ur)

    # 4. 路由规则 — FK columns: 关联上游(ID)=col 8, 插件组=col 9
    unm = data["up_name_map"]
    pm = data["plugins_map"]
    pcum = data["pc_uuid_map"]
    rr = []
    for r in data["routes"]:
        un = unm.get(r.upstream_id, "") if r.upstream_id else ""
        pls = pm.get(r.id, [])
        ps = "; ".join(
            f"{p.plugin_name}: {_fmt_json(p.config)}"
            for p in pls) if pls else "（无）"

        # FK: 关联上游(ID) → col 8 (0-based) — hyperlink to 上游服务!A{row}
        if r.upstream_id and r.upstream_id in up_row:
            uid_link = (r.upstream_id, "上游服务", f"A{up_row[r.upstream_id]}")
        else:
            uid_link = ""

        # FK: 插件组 → col 9 (0-based) — resolve edge_uuids to hyperlinks
        pc_ids_str = r.plugin_config_ids or ""
        resolved_pcs = ""
        if pc_ids_str:
            try:
                uuids = json.loads(pc_ids_str)
                parts = []
                for uuid in uuids:
                    info = pcum.get(uuid)
                    if info and uuid in pc_row:
                        link = f"{info[1]} (id: {info[0]})"
                        parts.append((link, "插件组", f"A{pc_row[uuid]}"))
                    elif info:
                        parts.append(f"{info[1]} (id: {info[0]})")
                    else:
                        parts.append(f"{uuid} (未匹配)")
                if parts and all(isinstance(p, tuple) for p in parts):
                    resolved_pcs = "; ".join(p[0] for p in parts)
                    # Store as single string with multiple links — keep as string for now
                    resolved_pcs = "; ".join(
                        p[0] if isinstance(p, tuple) else p for p in parts
                    )
                else:
                    resolved_pcs = "; ".join(
                        p[0] if isinstance(p, tuple) else p for p in parts
                    )
            except (json.JSONDecodeError, TypeError):
                resolved_pcs = pc_ids_str

        rr.append([
            r.id, r.name, r.uri or "", r.methods or "", r.hosts or "",
            r.priority, "启用" if r.status == 1 else "禁用",
            un, uid_link, resolved_pcs, ps,
            r.description or "", _fmt_dt(r.created_at),
        ])
    write_sheet("路由规则", [
        "ID", "名称", "URI", "方法", "域名", "优先级", "状态",
        "关联上游(名称)", "关联上游(ID)", "插件组", "插件", "描述", "创建时间",
    ], rr, link_cols={8})

    # 5. 插件组
    pcr = []
    for p in data["plugin_configs"]:
        pcr.append([
            p.id, p.name, _fmt_json(p.plugins), p.description or "",
            _fmt_dt(p.created_at),
        ])
    write_sheet("插件组", [
        "ID", "名称", "插件配置", "描述", "创建时间",
    ], pcr)

    # 6. 全局规则
    grr = []
    for g in data["global_rules"]:
        grr.append([
            g.id, g.name, _fmt_json(g.plugins), g.description or "",
            _fmt_dt(g.created_at),
        ])
    write_sheet("全局规则", [
        "ID", "名称", "插件配置", "描述", "创建时间",
    ], grr)

    # 7. 插件元数据
    pmr = []
    for p in data["plugin_metadatas"]:
        pmr.append([
            p.id, p.plugin_name, _fmt_json(p.config_data),
            _fmt_dt(p.created_at),
        ])
    write_sheet("插件元数据", [
        "ID", "插件名称", "配置数据", "创建时间",
    ], pmr)

    # 8. 四层代理
    spr = []
    for s in data["stream_proxies"]:
        spr.append([
            s.id, s.name, s.listen_port, s.scheme or "", s.load_balance or "",
            _fmt_json(s.targets), s.proxy_type,
            "启用" if s.status == 1 else "禁用",
            s.description or "", _fmt_dt(s.created_at),
        ])
    write_sheet("四层代理", [
        "ID", "名称", "监听端口", "协议", "负载均衡",
        "目标节点", "代理类型", "状态", "描述", "创建时间",
    ], spr)

    # 9. 静态资源 — FK column: 关联路由(ID)=col 5 (0-based)
    rnm = data["rt_name_map"]
    srr = []
    for s in data["static_resources"]:
        rn = rnm.get(s.route_id, "") if s.route_id else ""
        if s.route_id and s.route_id in rt_row:
            rid_link = (s.route_id, "路由规则", f"A{rt_row[s.route_id]}")
        else:
            rid_link = ""
        srr.append([
            s.id, s.name, s.url_path or "", s.file_size or 0,
            rn, rid_link,
            s.description or "", _fmt_dt(s.created_at),
        ])
    write_sheet("静态资源", [
        "ID", "名称", "URL 路径", "文件大小(字节)",
        "关联路由(名称)", "关联路由(ID)", "描述", "创建时间",
    ], srr, link_cols={5})

    # 10. SSL 证书 (metadata only - excluding cert, private_key, sign_cert, sign_key, client_ca, generate_log)
    slr = []
    for s in data["ssl_certificates"]:
        slr.append([
            s.id, s.name, s.sni, s.cert_type, s.algorithm or "",
            s.organization or "", "是" if s.is_ca else "否",
            s.create_method, "启用" if s.status == 1 else "禁用",
            _fmt_dt(s.created_at),
        ])
    write_sheet("SSL 证书", [
        "ID", "名称", "SNI 域名", "证书类型", "算法",
        "组织", "CA 证书", "创建方式", "状态", "创建时间",
    ], slr)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@router.get("/{cluster_id}/export")
async def export_cluster_data(
    cluster_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    data = await _query_all_data(cluster_id, db)
    buf = _build_workbook(data)

    cluster_name = data["cluster"].name
    filename = f"{_sanitize_filename(cluster_name)}_配置导出.xlsx"

    from urllib.parse import quote

    encoded = quote(filename, safe="")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                f"attachment; filename*=UTF-8''{encoded}"
            ),
        },
    )
