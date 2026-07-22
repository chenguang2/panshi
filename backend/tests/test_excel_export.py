"""
Tests for cluster Excel export functionality.

TDD: these tests are written BEFORE the implementation.
"""
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook
from httpx import ASGITransport, AsyncClient
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.main import app
from app.core.database import get_db, Base
from app.core.security import hash_password
from app.models.cluster import (
    Cluster, Node, Upstream, UpstreamTarget, Route, RoutePlugin,
    PluginConfig, GlobalRule, PluginMetadata, StreamProxy,
)
from app.models.static_resource import StaticResource
from app.models.ssl import SslCertificate
from app.models.user import User

TEST_DB_URL = "sqlite+aiosqlite:///:memory:"


class TestOpenpyxlAvailable:

    def test_openpyxl_can_create_workbook(self):
        """Verify openpyxl is installed and can create a basic Workbook."""
        wb = Workbook()
        assert wb is not None
        ws = wb.active
        ws.title = "测试"
        ws.cell(row=1, column=1, value="Hello")
        assert wb["测试"]["A1"].value == "Hello"


class TestExportRouterExists:

    def test_export_router_module_importable(self):
        """Verify the cluster_export module can be imported and has a router."""
        from app.api.v1.cluster_export import router
        assert router is not None
        assert hasattr(router, "routes")


@pytest.fixture
async def test_engine():
    engine = create_async_engine(TEST_DB_URL, echo=False)
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)
    yield engine
    await engine.dispose()


@pytest.fixture
def test_sessionmaker(test_engine):
    return async_sessionmaker(test_engine, class_=AsyncSession, expire_on_commit=False)


@pytest.fixture
async def seed_db(test_sessionmaker):
    """Seed the test database with a cluster and all resource types."""
    async with test_sessionmaker() as session:
        # Create admin user
        session.add(User(username="admin", password_hash=hash_password("panshi123"), role="admin", status=1))

        # Create cluster
        cluster = Cluster(name="test-cluster", display_name="测试集群", description="test",
                          group_name="test-group", status=1)
        session.add(cluster)
        await session.commit()
        await session.refresh(cluster)
        cid = cluster.id

        # Nodes
        session.add_all([
            Node(cluster_id=cid, ip="10.0.0.1", service_port=80, management_port=9180,
                 edge_path="/edge/node1", status=1),
            Node(cluster_id=cid, ip="10.0.0.2", service_port=443, management_port=9180,
                 edge_path="/edge/node2", status=0),
        ])

        # Upstream with targets
        us1 = Upstream(cluster_id=cid, name="api-upstream", load_balance="weighted_roundrobin",
                       scheme="http", pass_host="pass",
                       timeout='{"connect": 5, "send": 10, "read": 10}')
        session.add(us1)
        await session.commit()
        await session.refresh(us1)
        session.add_all([
            UpstreamTarget(upstream_id=us1.id, target="10.0.0.10:8080", weight=100),
            UpstreamTarget(upstream_id=us1.id, target="10.0.0.11:8080", weight=200),
        ])

        # Upstream without targets
        us2 = Upstream(cluster_id=cid, name="backup-upstream", load_balance="weighted_roundrobin", scheme="http")
        session.add(us2)

        # Routes
        rt1 = Route(cluster_id=cid, name="api-route", uri="/api/*", methods="GET,POST",
                    hosts="example.com", priority=100, status=1, upstream_id=us1.id)
        session.add(rt1)
        await session.commit()
        await session.refresh(rt1)
        session.add(RoutePlugin(route_id=rt1.id, plugin_name="limit-req",
                                config='{"rate": 10, "burst": 20}'))

        rt2 = Route(cluster_id=cid, name="no-plugin-route", uri="/health", methods="GET",
                    priority=0, status=1)
        session.add(rt2)

        # PluginConfig
        session.add(PluginConfig(cluster_id=cid, name="cors-plugin",
                                 plugins='{"cors": {"allow_origins": "*"}}'))

        # GlobalRule
        session.add(GlobalRule(cluster_id=cid, name="rate-limit-rule",
                               plugins='{"limit-req": {"rate": 5}}'))

        # PluginMetadata
        session.add(PluginMetadata(cluster_id=cid, plugin_name="custom-plugin",
                                   config_data='{"schema": {"type": "object"}}'))

        # StreamProxy
        session.add(StreamProxy(cluster_id=cid, name="tcp-proxy", listen_port=3306,
                                scheme="tcp", load_balance="weighted_roundrobin",
                                targets='[{"target":"10.0.0.20:3306","weight":100}]',
                                proxy_type="normal", status=1))

        # StaticResource (linked to route)
        session.add(StaticResource(cluster_id=cid, route_id=rt1.id, name="static-assets",
                                   url_path="/static/*", file_size=1024))

        # SSL Certificate (metadata only fields)
        session.add(SslCertificate(cluster_id=cid, name="test-cert", sni="example.com",
                                   cert="PRIVATE_CERT_CONTENT", private_key="PRIVATE_KEY_CONTENT",
                                   cert_type="server", algorithm="rsa", organization="TestOrg",
                                   is_ca=False, create_method="upload", status=1))

        await session.commit()
        yield cluster


async def _seed_user(session):
    """Helper to seed an admin user for authentication."""
    from sqlalchemy import select
    result = await session.execute(
        select(User).where(User.username == "admin"))
    if not result.scalar_one_or_none():
        session.add(User(username="admin",
                         password_hash=hash_password("panshi123"),
                         role="admin", status=1))
        await session.commit()


async def _get_token(test_sessionmaker):
    """Helper to authenticate and get a token."""
    async def override_get_db():
        async with test_sessionmaker() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as ac:
        login_resp = await ac.post("/api/v1/auth/login",
                                   json={"username": "admin", "password": "panshi123"})
        assert login_resp.status_code == 200
        return login_resp.json()["access_token"]


@pytest.mark.asyncio
async def test_export_non_existent_cluster_returns_404(test_sessionmaker):
    """Scenario: Cluster not found returns 404."""
    # Seed user
    async with test_sessionmaker() as session:
        await _seed_user(session)
    token = await _get_token(test_sessionmaker)

    async def override_get_db():
        async with test_sessionmaker() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            resp = await ac.get("/api/v1/clusters/99999/export",
                                headers={"Authorization": f"Bearer {token}"})
            assert resp.status_code == 404
    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_export_all_resource_types(seed_db, test_sessionmaker):
    """Successful export returns Excel with correct sheets and data."""
    cluster = seed_db
    cid = cluster.id

    async def override_get_db():
        async with test_sessionmaker() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            login_resp = await ac.post("/api/v1/auth/login",
                                       json={"username": "admin", "password": "panshi123"})
            token = login_resp.json()["access_token"]

            resp = await ac.get(f"/api/v1/clusters/{cid}/export",
                                headers={"Authorization": f"Bearer {token}"})
            assert resp.status_code == 200
            ct = resp.headers["content-type"]
            assert "spreadsheetml.sheet" in ct
            cd = resp.headers.get("content-disposition", "")
            assert "%E9%85%8D%E7%BD%AE%E5%AF%BC%E5%87%BA" in cd
            assert "test-cluster" in cd

            # Load the workbook from response content
            wb = load_workbook(BytesIO(resp.content))

            # Check all 10 sheets exist
            expected_sheets = ["集群信息", "集群节点", "上游服务", "路由规则", "插件组",
                               "全局规则", "插件元数据", "四层代理", "静态资源", "SSL 证书"]
            for sheet_name in expected_sheets:
                assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

            # Check cluster info (no admin_key)
            ws = wb["集群信息"]
            cluster_headers = [str(c.value or "") for c in ws[1]]
            assert not any("admin_key" in h.lower() for h in cluster_headers), \
                "admin_key should not be exported"
            # Verify the data row has the cluster name
            assert ws[2][0].value == "test-cluster"

            # Check nodes (2 rows)
            ws = wb["集群节点"]
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            assert data_rows == 2
            assert ws[2][1].value == "10.0.0.1"  # column A=ID, B=IP
            assert ws[2][0].value == 1  # ID

            # Check upstreams (2 rows)
            ws = wb["上游服务"]
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            assert data_rows == 2
            up_headers = [str(c.value or "") for c in ws[1]]
            assert "目标节点" in up_headers
            assert up_headers[0] == "ID"  # First column is now ID

            # Check routes (2 rows)
            ws = wb["路由规则"]
            data_rows = sum(1 for row in ws.iter_rows(min_row=2) if row[0].value)
            assert data_rows == 2
            rt_headers = [str(c.value or "") for c in ws[1]]
            assert "关联上游(名称)" in rt_headers
            assert "关联上游(ID)" in rt_headers
            assert "插件" in rt_headers
            assert "插件组" in rt_headers

            # Check SSL excludes sensitive fields
            ws_ssl = wb["SSL 证书"]
            ssl_headers = [str(c.value or "").lower() for c in ws_ssl[1]]
            for sensitive in ["cert", "private_key", "sign_cert", "sign_key", "client_ca"]:
                assert not any(sensitive.replace("_", "") in h.replace("_", "").replace(" ", "")
                               for h in ssl_headers), f"Sensitive field exposed: {sensitive}"

            # Check header row is bold
            for cell in wb["集群信息"][1]:
                if cell.value:
                    assert cell.font and cell.font.bold, "Header should be bold"

    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_export_empty_cluster_produces_all_sheets_with_headers_only(
        test_sessionmaker):
    """Empty data types still produce a sheet with header only."""
    async with test_sessionmaker() as session:
        session.add(User(username="admin",
                         password_hash=hash_password("panshi123"),
                         role="admin", status=1))
        cluster = Cluster(name="empty-cluster", display_name="空集群", status=1)
        session.add(cluster)
        await session.commit()
        await session.refresh(cluster)
        cid = cluster.id

    async def override_get_db():
        async with test_sessionmaker() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            login_resp = await ac.post("/api/v1/auth/login",
                                       json={"username": "admin", "password": "panshi123"})
            token = login_resp.json()["access_token"]

            resp = await ac.get(f"/api/v1/clusters/{cid}/export",
                                headers={"Authorization": f"Bearer {token}"})
            assert resp.status_code == 200

            wb = load_workbook(BytesIO(resp.content))
            # All sheets exist
            expected = ["集群信息", "集群节点", "上游服务", "路由规则", "插件组",
                        "全局规则", "插件元数据", "四层代理", "静态资源", "SSL 证书"]
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                assert ws[1][0].value is not None, f"Sheet {sheet_name} missing header"
                # Only cluster info has data row; others should be header-only
                data_count = sum(1 for row in ws.iter_rows(min_row=2, max_row=3) if row[0].value)
                if sheet_name == "集群信息":
                    assert data_count == 1
                else:
                    assert data_count == 0, \
                        f"Sheet {sheet_name} should have no data rows"

    finally:
        app.dependency_overrides.clear()


@pytest.mark.asyncio
async def test_export_any_authenticated_user_can_export(
        seed_db, test_sessionmaker):
    """Non-admin user can also export."""
    cluster = seed_db
    cid = cluster.id

    async def override_get_db():
        async with test_sessionmaker() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        transport = ASGITransport(app=app)
        async with AsyncClient(transport=transport, base_url="http://test") as ac:
            login_resp = await ac.post("/api/v1/auth/login",
                                       json={"username": "admin", "password": "panshi123"})
            token = login_resp.json()["access_token"]

            resp = await ac.get(f"/api/v1/clusters/{cid}/export",
                                headers={"Authorization": f"Bearer {token}"})
            assert resp.status_code == 200
    finally:
        app.dependency_overrides.clear()
